# -*- coding:UTF-8 -*- #
"""
@filename:Hierarchical_clustering_export_final_with_print.py
@author:Weaki
@time:2025-10-08
"""
import datetime

import pandas as pd
from typing import List, Dict, Any
from collections import defaultdict
import time
# 导入用于Excel格式化的库
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import sys
import io
import os
import re
from gooey import Gooey, GooeyParser
# --- 强制标准输出/错误流使用 UTF-8 编码并启用行缓冲 ---
# 这是一个处理打包后程序（尤其是在Windows上）Unicode错误和输出延迟问题的稳定方法。
# line_buffering=True 确保每行 print 输出后都会立即刷新，实现实时显示。
if sys.stdout is not None:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)
if sys.stderr is not None:
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', line_buffering=True)


class HierarchicalCluster:
    """
    Excel数据的多级层次聚类工具

    特点:
    - 支持任意多级分层
    - 只存储行索引,节省内存
    - 自动检测空子表格
    - 递归构建完整层次结构
    - [最终版] 支持按层级导出到Excel，实现正确排序、单元格合并、添加计数和居中格式化
    """

    def __init__(self, df: pd.DataFrame, columns: List[str], complete_mode: bool = False):
        self.df = df
        self.columns = columns
        self.complete_mode = complete_mode
        self.cluster_tree = {}
        self.empty_clusters = []
        if complete_mode:
            self.all_values = {col: sorted(df[col].unique().tolist()) for col in columns}

    def cluster(self) -> Dict[str, Any]:
        start_time = time.time()
        all_indices = list(self.df.index)
        self.cluster_tree = self._build_cluster_tree(indices=all_indices, level=0, path=[])
        elapsed_time = time.time() - start_time
        return {
            'tree': self.cluster_tree, 'empty_clusters': self.empty_clusters,
            'stats': {
                'total_rows': len(self.df), 'levels': len(self.columns),
                'empty_count': len(self.empty_clusters), 'time_elapsed': f"{elapsed_time:.4f}秒"
            }
        }

    def _build_cluster_tree(self, indices: List[int], level: int, path: List[str]) -> Dict:
        if level >= len(self.columns): return {}
        current_col = self.columns[level]
        cluster_dict = {}
        if self.complete_mode:
            values_to_process = self.all_values[current_col]
        else:
            if len(indices) == 0:
                return {}
            values_to_process = sorted(self.df.loc[indices, current_col].unique())
        groups = defaultdict(list)
        for idx in indices:
            value = self.df.loc[idx, current_col]
            groups[value].append(idx)
        for value in values_to_process:
            group_indices = groups.get(value, [])
            current_path = path + [str(value)]
            path_str = '->'.join(current_path)
            if len(group_indices) == 0 and self.complete_mode: self.empty_clusters.append(path_str)
            node_info = {
                'path': path_str, 'level': level, 'field': current_col, 'value': value,
                'row_indices': group_indices, 'row_count': len(group_indices),
                'children': {}, 'is_empty': len(group_indices) == 0
            }
            if level < len(self.columns) - 1:
                node_info['children'] = self._build_cluster_tree(indices=group_indices, level=level + 1,
                                                                 path=current_path)
            cluster_dict[value] = node_info
        return cluster_dict

    # =================================================================
    # 🌟 新增：打印树结构的方法
    # =================================================================
    def print_tree(self, max_depth: int = None):
        """
        打印聚类树结构

        Args:
            max_depth: 最大显示深度,None表示显示全部
        """
        print("\n" + "=" * 60)
        print("🌳 聚类树结构")
        print("=" * 60)
        if not self.cluster_tree:
            print("树为空，请先运行 .cluster() 方法。")
            return

        self._print_node(self.cluster_tree, 0, max_depth)

        if self.empty_clusters:
            print("\n" + "=" * 60)
            print(f"空子表格警告 (共{len(self.empty_clusters)}个):")
            print("=" * 60)
            for path in self.empty_clusters:
                print(f"  ⚠ {path}")

    def _print_node(self, node: Dict, depth: int, max_depth: int):
        """递归打印节点"""
        if max_depth is not None and depth >= max_depth:
            return

        for key, info in sorted(node.items()):
            indent = "  " * depth
            path = info.get('path', str(key))
            count = info.get('row_count', 0)
            is_empty = info.get('is_empty', False)

            # 空节点用特殊标记
            empty_mark = " ⚠️ [空]" if is_empty else ""
            print(f"{indent}├─ {key} ({count}行){empty_mark}")

            if 'children' in info and info['children']:
                self._print_node(info['children'], depth + 1, max_depth)

    def get_level_indices(self, level: int) -> Dict[str, List[int]]:
        if level < 0 or level >= len(self.columns):
            raise ValueError(f"层级必须在 0 到 {len(self.columns) - 1} 之间")
        result = {}
        self._collect_level_nodes(self.cluster_tree, 0, level, result)
        return result

    def _get_all_node_counts(self) -> Dict[str, int]:
        """
        递归遍历整个树，获取所有节点的完整路径及其计数。

        Returns:
            Dict[str, int]: { '路径' -> 数量 } 的字典
        """
        counts = {}

        def traverse(node_dict):
            """递归遍历"""
            for key, info in node_dict.items():
                # 存储当前节点的路径和计数
                counts[info['path']] = info['row_count']
                # 如果有子节点，继续遍历
                if info.get('children'):
                    traverse(info['children'])

        # 从树的根节点开始遍历
        traverse(self.cluster_tree)
        return counts

    def _collect_level_nodes(self, node: Dict, current_level: int, target_level: int, result: Dict):
        for key, info in node.items():
            if current_level == target_level:
                result[info['path']] = info['row_indices']
            elif current_level < target_level and info.get('children'):
                self._collect_level_nodes(info['children'], current_level + 1, target_level, result)

    def _format_and_merge_sheet(self, worksheet, merge_cols_indices: List[int]):
        """
        对给定的worksheet进行格式化：居中、合并、调整列宽
        """
        if not worksheet:
            return

        center_align = Alignment(horizontal='center', vertical='center')

        # 1. 居中所有单元格
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = center_align

        # 2. 合并指定列的单元格
        for col_idx in merge_cols_indices:
            start_row = 2
            for i in range(2, worksheet.max_row + 2):
                is_last_row = (i == worksheet.max_row + 1)
                current_cell_value = worksheet.cell(row=i, column=col_idx + 1).value
                start_cell_value = worksheet.cell(row=start_row, column=col_idx + 1).value

                if is_last_row or (current_cell_value != start_cell_value):
                    if i - 1 > start_row:
                        worksheet.merge_cells(start_row=start_row, start_column=col_idx + 1,
                                              end_row=i - 1, end_column=col_idx + 1)
                    start_row = i

        # 3. 自动调整列宽
        for col in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    # 将表头长度也纳入计算
                    header_len = len(str(worksheet.cell(row=1, column=cell.column).value))
                    cell_len = len(str(cell.value))
                    current_max = max(header_len, cell_len)
                    if current_max > max_length:
                        max_length = current_max
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def export_to_excel_by_level(self, output_path: str):
        if not self.cluster_tree:
            print("❌ 错误: 请先调用 .cluster() 方法执行聚类。")
            return
        try:
            print(f"\n📦 正在导出(明细版)Excel到: {output_path}")
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                all_path_counts = {}
                # 一次性获取所有计数
                all_path_counts = self._get_all_node_counts()

                # --- Sheet 1: 第一级 ---
                if len(self.columns) >= 1:
                    level_0_indices = self.get_level_indices(0)
                    level_0_data = []
                    for path, indices in level_0_indices.items():
                        count = len(indices)
                        # all_path_counts[path] = count
                        level_0_data.append({self.columns[0]: path, '数量': count})
                    if level_0_data:
                        df_level_0 = pd.DataFrame(level_0_data).sort_values(by=['数量', self.columns[0]],
                                                                            ascending=[False, True])
                        df_level_0.to_excel(writer, sheet_name='第一级汇总', index=False)
                        self._format_and_merge_sheet(writer.sheets.get('第一级汇总'), merge_cols_indices=[])

                # --- Sheet 2: 第二级 ---
                if len(self.columns) >= 2:
                    level_1_indices = self.get_level_indices(1)
                    level_1_data = []
                    for path, indices in level_1_indices.items():
                        count = len(indices)
                        # all_path_counts[path] = count
                        parts = path.split('->')
                        parent_path = parts[0]
                        level_1_data.append({
                            self.columns[0]: parts[0], self.columns[1]: parts[1], '数量': count,
                            '__parent_count': all_path_counts.get(parent_path, 0)
                        })
                    if level_1_data:
                        df_level_1 = pd.DataFrame(level_1_data).sort_values(
                            by=['__parent_count', self.columns[0], '数量'], ascending=[False, True, False]
                        )
                        df_level_1[self.columns[0]] = df_level_1.apply(
                            lambda r: f"{r[self.columns[0]]} ({r['__parent_count']})", axis=1)
                        df_level_1.drop(columns=['__parent_count']).to_excel(writer, sheet_name='第二级汇总',
                                                                             index=False)
                        self._format_and_merge_sheet(writer.sheets.get('第二级汇总'), merge_cols_indices=[0])

                # --- Sheet 3: 第三级 ---
                if len(self.columns) >= 3:
                    level_2_indices = self.get_level_indices(2)
                    level_2_data = []
                    for path, indices in level_2_indices.items():
                        count = len(indices)
                        parts = path.split('->')
                        parent_path_l1 = parts[0]
                        parent_path_l2 = '->'.join(parts[:2])
                        level_2_data.append({
                            self.columns[0]: parts[0], self.columns[1]: parts[1], self.columns[2]: parts[2],
                            '数量': count,
                            '__parent_count_l1': all_path_counts.get(parent_path_l1, 0),
                            '__parent_count_l2': all_path_counts.get(parent_path_l2, 0)
                        })
                    if level_2_data:
                        df_level_2 = pd.DataFrame(level_2_data).sort_values(
                            by=['__parent_count_l1', self.columns[0], '__parent_count_l2', self.columns[1], '数量'],
                            ascending=[False, True, False, True, False]
                        )
                        df_level_2[self.columns[0]] = df_level_2.apply(
                            lambda r: f"{r[self.columns[0]]} ({r['__parent_count_l1']})", axis=1)
                        df_level_2[self.columns[1]] = df_level_2.apply(
                            lambda r: f"{r[self.columns[1]]} ({r['__parent_count_l2']})", axis=1)
                        df_level_2.drop(columns=['__parent_count_l1', '__parent_count_l2']).to_excel(writer,
                                                                                                     sheet_name='第三级汇总',
                                                                                                     index=False)
                        self._format_and_merge_sheet(writer.sheets.get('第三级汇总'), merge_cols_indices=[0, 1])

            print(f"\n✅ 成功导出明细版格式化Excel文件到: {output_path}")
        except Exception as e:
            print(f"\n❌ 导出Excel失败: {e}")
            print("  请确保您已安装 'openpyxl' 库 (在终端或命令提示符中运行: pip install openpyxl)")

    def export_to_excel_aggregated(self, output_path: str, threshold_percent: float = 0.05):
        """
        [新功能] 以聚合方式导出到Excel。
        最后一级将作为聚合字符串（带阈值）显示在单元格中。

        Args:
            output_path (str): 导出路径。
            threshold_percent (float): 介于0.0到1.0之间。
                子类别占父类别总数的百分比，必须 **大于等于** 此阈值才会被显示。
                例如 0.05 = 5%。
        """
        if not self.cluster_tree:
            print("❌ 错误: 请先调用 .cluster() 方法执行聚类。")
            return

        print(f"📦 正在导出(聚合版)Excel到: {output_path}")
        if threshold_percent > 0:
            print(f"  (将隐藏占比 < {threshold_percent:.1%} 的子类别)")
        else:
            print("  (threshold_percent=0, 将显示所有子类别)")

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

                # 1. 一次性获取所有节点的计数
                all_path_counts = self._get_all_node_counts()

                # 2. 遍历每个父层级 (例如 3层['A','B','C']，循环 0 和 1)
                for parent_level in range(len(self.columns) - 1):
                    child_level = parent_level + 1
                    child_col_name = self.columns[child_level]
                    sheet_name = f'第{child_level}级聚合'  # "第2级聚合", "第3级聚合"

                    # 3. 递归收集所有父层级的节点
                    parent_nodes = []

                    def collect_nodes(node_dict, current_level):
                        """递归查找所有在 parent_level 上的节点"""
                        if current_level == parent_level:
                            for info in node_dict.values():
                                parent_nodes.append(info)
                            return  # 找到即停止，不再深入

                        # 如果没到层级，继续深入
                        for info in node_dict.values():
                            if info.get('children'):
                                collect_nodes(info['children'], current_level + 1)

                    collect_nodes(self.cluster_tree, 0)

                    # 4. 构建数据
                    agg_data = []

                    for parent_node in parent_nodes:
                        parent_path_parts = parent_node['path'].split('->')
                        parent_total_count = parent_node['row_count']

                        # 如果父节点为空，则跳过
                        if parent_total_count == 0:
                            continue

                        # 创建基础行 (e.g., {'A': 'A1', 'B': 'B1'})
                        row = {self.columns[i]: part for i, part in enumerate(parent_path_parts)}

                        # 5. 处理子节点，进行聚合
                        child_nodes = parent_node.get('children', {}).values()

                        # 按数量降序排序子节点
                        sorted_children = sorted(child_nodes, key=lambda x: x['row_count'], reverse=True)

                        agg_strings = []
                        for child_node in sorted_children:
                            child_count = child_node['row_count']
                            if child_count == 0:
                                continue  # 跳过空的子节点

                            percentage = child_count / parent_total_count

                            # 检查阈值
                            if percentage >= threshold_percent:
                                # 格式: C1(4)
                                agg_strings.append(f"{child_node['value']}({child_count})")

                        # 6. 组合成单元格内容，使用中文顿号 '、'
                        row[f'{child_col_name}_聚合'] = '、'.join(agg_strings)
                        row['总数'] = parent_total_count

                        # 7. 添加用于排序的父级计数
                        if parent_level > 0:  # 仅 L2->L3 需要 L1 的计数
                            l1_path = parent_path_parts[0]
                            row['__parent_count_l1'] = all_path_counts.get(l1_path, 0)

                        agg_data.append(row)

                    if not agg_data:
                        continue  # 如果没数据，跳到下一级

                    # 8. 转换为DataFrame并排序、格式化
                    df_agg = pd.DataFrame(agg_data)

                    sort_cols = []
                    sort_asc = []
                    merge_cols = []

                    # 排序逻辑 (模仿原函数)
                    if parent_level == 0:  # L1 -> L2 (A -> B聚合)
                        sort_cols = ['总数', self.columns[0]]
                        sort_asc = [False, True]

                    elif parent_level == 1:  # L2 -> L3 (A, B -> C聚合)
                        sort_cols = ['__parent_count_l1', self.columns[0], '总数', self.columns[1]]
                        sort_asc = [False, True, False, True]  # L1_count(D), L1_name(A), L2_total(D), L2_name(A)

                        df_agg = df_agg.sort_values(by=sort_cols, ascending=sort_asc)

                        # 在L1列的值后面加上L1的总数
                        df_agg[self.columns[0]] = df_agg.apply(
                            lambda r: f"{r[self.columns[0]]} ({r['__parent_count_l1']})", axis=1)

                        df_agg = df_agg.drop(columns=['__parent_count_l1'])
                        merge_cols = [0]  # 合并L1 (第0列)

                    # L1->L2 的排序
                    if parent_level == 0:
                        df_agg = df_agg.sort_values(by=sort_cols, ascending=sort_asc)

                    # 9. 写入Excel
                    # 重新排列表头顺序
                    final_cols = self.columns[:child_level] + [f'{child_col_name}_聚合', '总数']
                    df_agg = df_agg[final_cols]

                    df_agg.to_excel(writer, sheet_name=sheet_name, index=False)

                    # 10. 格式化
                    self._format_and_merge_sheet(writer.sheets.get(sheet_name), merge_cols_indices=merge_cols)

            print(f"\n✅ 成功导出聚合版Excel文件到: {output_path}")
        except PermissionError:
            print(f"\n❌ 导出Excel失败: 权限错误。请关闭正在打开的 '{output_path}' 文件后再试！")
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"\n❌ 导出聚合Excel失败: {e}")

# # ============ 使用示例 ============
#
# def example_from_excel(file_path: str, columns: List[str]):
#     """从Excel文件读取并聚类的示例"""
#     try:
#         df = pd.read_excel(file_path)
#         print(f"✅ 成功从 {file_path} 读取 {len(df)} 行数据。")
#
#         # 1. 初始化聚类器
#         clusterer = HierarchicalCluster(df, columns=columns)
#
#         # 2. 执行聚类
#         clusterer.cluster()
#
#         # 3. 打印聚类树
#         clusterer.print_tree()
#
#         # 4. 导出到Excel (两种方式任选)
#
#         # 方式一：原始明细版
#         output_filename = "Excel聚类结果(明细版).xlsx"
#         clusterer.export_to_excel_by_level(output_filename)
#
#         # 🌟 方式二：新的聚合版
#         output_agg_filename = "Excel聚类结果(聚合版).xlsx"
#         # 阈值设为 0.1 (即 10%)，占比 <= 10% 的子类将被隐藏
#         clusterer.export_to_excel_aggregated(output_agg_filename, threshold_percent=0.5)
#
#         # # 如果想显示所有，可以设为 0
#         # clusterer.export_to_excel_aggregated(output_agg_filename, threshold_percent=0)
#
#     except FileNotFoundError:
#         print(f"❌ 错误: 文件未找到 - {file_path}")
#         print("  请检查文件路径是否正确，特别是路径中的斜杠。")
#     except KeyError as e:
#         print(f"❌ 错误: 列名 {e} 不存在。")
#         print(f"  请确保您的Excel文件中包含以下所有列: {columns}")
#     except Exception as e:
#         print(f"❌ 读取或处理Excel时发生错误: {e}")
#
#
# if __name__ == "__main__":
#     print("\n" + "=" * 30, "运行Excel文件示例", "=" * 30)
#     # ⚠️ 请确保您的桌面上有这个 test.xlsx 文件，或者修改为您的正确路径
#     excel_file = r"C:\Users\weaki\Desktop\test.xlsx"
#     # ⚠️ 请修改为您的分组列名
#     group_columns = ['A', 'B', 'C']
#     example_from_excel(excel_file, group_columns)

# ============ Gooey 调用的核心逻辑 ============

def run_clustering_logic(args, columns_list):
    """
    Gooey 调用的核心逻辑函数
    (这个函数替代了旧的 example_from_excel)
    """
    try:
        df = pd.read_excel(args.input_file)
        print(f"✅ 成功从 {args.input_file} 读取 {len(df)} 行数据。")

        # 1. 初始化聚类器 (使用来自GUI的 'complete_mode')
        clusterer = HierarchicalCluster(df, columns=columns_list, complete_mode=args.complete_mode)

        # 2. 执行聚类
        clusterer.cluster()

        # 3. 打印聚类树 (会显示在Gooey的控制台)
        clusterer.print_tree()

        # 4. 导出到Excel (使用来自GUI的 'output_directory')

        # 方式一：原始明细版
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_detailed = os.path.join(args.output_directory, f"Excel聚类结果(明细版)_{timestamp}.xlsx")
        clusterer.export_to_excel_by_level(output_detailed)

        # 方式二：新的聚合版 (使用来自GUI的 'threshold')
        output_agg = os.path.join(args.output_directory, f"Excel聚类结果(聚合版)_{timestamp}.xlsx")

        # 从GUI的百分比 (0-100) 转换为小数 (0.0-1.0)
        threshold_percent = args.threshold / 100.0

        clusterer.export_to_excel_aggregated(output_agg, threshold_percent=threshold_percent)

        print(f"\n" + "=" * 60)
        print(f"🎉 全部完成! 两个报告已保存至: {args.output_directory}")
        print("=" * 60)

    except FileNotFoundError:
        print(f"❌ 错误: 文件未找到 - {args.input_file}")
        print("  请检查文件路径是否正确。")
    except KeyError as e:
        print(f"❌ 错误: 列名 {e} 不存在。")
        print(f"  请确保您的Excel文件中包含以下所有列: {columns_list}")
    except Exception as e:
        print(f"❌ 读取或处理Excel时发生错误: {e}")
        import traceback
        traceback.print_exc()  # 向控制台打印详细错误


"""
分组统计
"""


def clean_sheet_name(name):
    """
        Excel Sheet名称不能包含特殊字符 : \ / ? * [ ]
        且长度不能超过31个字符。
    """
    if pd.isna(name):
        return "Unknown"
    # 将名称转换为字符串
    name = str(name)
    # 替换非法字符为下划线
    name = re.sub(r'[\\/*?:\[\]]', '_', name)
    # 截取前31个字符
    return name[:31]

def run_stats_logic(args):
    """ 执行分表统计逻辑 """
    file_path = args.stat_input_file
    group_col = args.stat_group_col
    target_col = args.stat_target_col

    print("=" * 60)
    print("📊 正在启动：Excel 分表统计工具")
    print(f"  源文件: {file_path}")
    print(f"  分表依据列: {group_col}")
    print(f"  统计目标列: {target_col}")
    print("=" * 60)
    try:
        # 读取Excel文件
        print(f"正在读取文件：{file_path}")
        df = pd.read_excel(file_path)
        # 检查列是否存在
        if group_col not in df.columns or target_col not in df.columns:
            print(f"错误: 列名 '{group_col}' 或 '{target_col}' 在文件中不存在。")
            return
        # 处理目标列的空值 (关键步骤)
        # 将 NaN 填充为 "空值"，确保统计时包含在内
        df[target_col] = df[target_col].fillna("空值")

        # 同样处理分表列的空值，防止分表时报错
        df[group_col] = df[group_col].fillna("未分类")
        # 准备输出文件名(增加时间戳放置覆盖)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"分组统计结果_{timestamp}.xlsx"
        # 创建ExcelWriter对象
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 获取分表列的所有唯一值
            unique_groups = df[group_col].unique()
            print(f"检测到 {group_col}列有{len(unique_groups)} 个分类，开始处理...")
            for group_val in unique_groups:
                # 筛选数据
                sub_df = df[df[group_col] == group_val]

                # 统计频次
                # value_counts 默认就是降序排列 (Descending)
                stats = sub_df[target_col].value_counts().reset_index()
                stats.columns = [target_col, '数量']

                # 计算百分比
                total_count = stats['数量'].sum()
                stats['百分比'] = (stats['数量']/total_count).apply(lambda x: f"{x:.2%}")
                # 写入Excel
                sheet_name = clean_sheet_name(group_val)
                # 将数据写入 Excel，从第 2 行开始写 (startrow=1)，给顶部标题留空间
                # index=False 不写入索引列
                stats.to_excel(writer, sheet_name=sheet_name, startrow=1, index=False)
                # --- C. 样式调整 (合并居中标题) ---

                # 获取当前 sheet 对象
                worksheet = writer.sheets[sheet_name]

                # 1. 设置顶部合并标题 (A1 到 C1)
                # 标题内容：显示分表列的名称和当前值，例如 "部门: 技术部"
                header_text = f"{group_col}: {group_val}"
                worksheet.merge_cells('A1:C1')  # 合并第一行的前三列
                cell_title = worksheet['A1']
                cell_title.value = header_text

                # 设置标题样式：居中、加粗、加大字号、背景色
                cell_title.alignment = Alignment(horizontal='center', vertical='center')
                cell_title.font = Font(bold=True, size=14, color="FFFFFF")
                cell_title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

                # 2. 调整列宽 (让每一列稍微宽一点，好看)
                worksheet.column_dimensions['A'].width = 25
                worksheet.column_dimensions['B'].width = 15
                worksheet.column_dimensions['C'].width = 15

                print(f"   ✅ 已生成分表: {sheet_name} (行数: {len(stats)})")

        print(f"\n🎉 全部完成！输出文件已保存为: {output_file}")
    except FileNotFoundError:
        print("错误: 找不到指定的文件，请检查路径。")
    except Exception as e:
        print(f"发生未知错误: {e}")
# ============ Gooey 界面定义 ============

@Gooey(
    program_name="Excel标签类表格处理工具",
    program_description="聚类/统计",
    navigation='TABBED',  # 关键设置：启用侧边栏/标签页模式
    default_size=(800, 600),
    language='chinese',  # 指定Gooey语言为中文
    encoding='UTF-8',  # 确保编码
    # terminal_font_color='#00FF00',  # 绿色控制台文字
    # terminal_panel_color='#333333'  # 深色控制台背景
)
def main():
    """
    Gooey的主函数，用于定义GUI界面
    """
    parser = GooeyParser(description="请选择左侧的功能模块进行操作")

    # 创建子解析器 (Subparsers)
    # dest='command' 用于后续判断用户选了哪个功能
    subs = parser.add_subparsers(help='功能列表', dest='command')
    # ========================================================
    # 功能 1: 聚类分析
    # ========================================================
    cluster_parser = subs.add_parser('Clustering', help='配置聚类选项')
    # --- 1. 输入设置 ---
    c_input_group = cluster_parser.add_argument_group("1. 输入设置", "选择源文件和分组列")
    c_input_group.add_argument(
        'input_file',
        metavar='Excel 源文件',
        help='请选择包含数据的Excel文件 (.xlsx, .xls)',
        widget='FileChooser',
        gooey_options={'wildcard': 'Excel 文件 (*.xlsx;*.xls)|*.xlsx;*.xls'}
    )
    c_input_group.add_argument(
        'group_columns',
        metavar='分组列名 (必填)',
        help='请按顺序输入要分组的列名，用英文逗号“,”隔开 (例如: A,B,C)',
        widget='TextField',
        # 添加验证，确保不为空
        gooey_options={
            'validator': {
                'test': 'user_input.strip() != ""',
                'message': '分组列名不能为空'
            }
        }
    )

    # --- 2. 输出设置 ---
    c_output_group = cluster_parser.add_argument_group("2. 输出设置", "选择报告保存位置")
    c_output_group.add_argument(
        'output_directory',
        metavar='报告保存目录',
        help='所有生成的Excel报告将保存在此文件夹中',
        widget='DirChooser'  # 目录选择器
    )

    # --- 3. 聚类选项 ---
    c_options_group = cluster_parser.add_argument_group("3. 聚类选项", "配置聚类和聚合报告的行为")
    c_options_group.add_argument(
        '--complete_mode',
        metavar='完整模式 (查漏)',
        help='勾选后，将分析所有可能的组合，并报告空缺项',
        action='store_true',  # 生成复选框
        default=False
    )
    c_options_group.add_argument(
        '--threshold',
        metavar='聚合阈值 (%)',
        help='在“聚合版”报告中，占比低于此百分比的子项将被隐藏',
        widget='Slider',  # 滑块
        gooey_options={'min': 0, 'max': 100, 'increment': 1},
        default=10,  # 默认 10%
        type=int  # 确保Gooey返回整数
    )

    # ========================================================
    # 功能 2: 分表统计
    # ========================================================
    stat_parser = subs.add_parser('Statistics', help='分表统计与占比')

    s_group = stat_parser.add_argument_group("分表统计设置", "根据某一列拆分Sheet并统计另一列的占比")

    s_group.add_argument(
        'stat_input_file',
        metavar='Excel 源文件',
        help='选择要统计的数据表',
        widget='FileChooser',
        gooey_options={'wildcard': 'Excel 文件 (*.xlsx;*.xls)|*.xlsx;*.xls'}
    )

    s_group.add_argument(
        'stat_group_col',
        metavar='分表列 (Group By)',
        help='将根据此列的不同值生成不同的Sheet',
        widget='TextField'
    )

    s_group.add_argument(
        'stat_target_col',
        metavar='统计列 (Count)',
        help='将统计此列在每个Sheet下的数量和百分比',
        widget='TextField'
    )

    # ========================================================
    # 解析与分发
    # ========================================================
    args = parser.parse_args()

    # 根据用户选择的子命令 (Clustering 或 Statistics) 分发到不同的逻辑函数
    if args.command == 'Clustering':
        # --- 4. 参数处理与逻辑调用 ---
        try:
            # 转换逗号分隔的字符串为列表
            columns_list = [col.strip() for col in args.group_columns.split(',') if col.strip()]
            if not columns_list:
                # 再次检查，以防Gooey验证器失效
                print("❌ 错误: “分组列名”不能为空，请至少输入一个列名。")
                return

            print("=" * 60)
            print("🚀 开始执行聚类... (请稍候，完成后会弹出提示)")
            print(f"  源文件: {args.input_file}")
            print(f"  分组列: {columns_list}")
            print(f"  保存目录: {args.output_directory}")
            print(f"  完整模式: {'是' if args.complete_mode else '否'}")
            print(f"  聚合阈值: {args.threshold}%")
            print("=" * 60)

            # 调用核心逻辑
            run_clustering_logic(args, columns_list)

        except Exception as e:
            print(f"❌ 发生致命错误: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)  # 退出并显示错误
    elif args.command == 'Statistics':
        run_stats_logic(args)
    else:
        print("请选择一个功能运行。")


# ============ 程序入口 ============
if __name__ == "__main__":
    main()
